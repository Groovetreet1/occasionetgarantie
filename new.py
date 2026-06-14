# -*- coding: utf-8 -*-
import streamlit as st
import pandas as pd
import zipfile
import re
import io
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
import os
import math

# ==============================
# CONFIG GLOBALE STREAMLIT
# ==============================
st.set_page_config(
    page_title="Suite Outils Téléphonie / Marketing",
    page_icon="https://media.licdn.com/dms/image/v2/D4E0BAQHd1vQ5srIY4w/company-logo_200_200/B4EZZdUUMCHMAI-/0/1745322330921/gomobile_africa_logo?e=2147483647&v=beta&t=h3LJTeBhImOortH_t5PmBxDMwzEi3vyIelylBx9lKuU",
    layout="wide"
)

# ==============================
# 1) APP EQDOM MARKETING
#    (fichier eqdom_mrkg_app.py)
# ==============================

DATE_EXPORT = datetime.now().strftime("%d%m%Y")  # ex : 21/11/2025 -> "21112025"


def traiter_classeur(uploaded_file):
    # Lecture de TOUTES les feuilles en texte
    xls = pd.read_excel(uploaded_file, sheet_name=None, dtype=str)

    processed_sheets = {}
    stats_rows = []

    for idx, (sheet_name, df) in enumerate(xls.items(), start=1):
        df_original = df.copy()
        total_initial = len(df_original)

        # Nettoyage des noms de colonnes
        df.columns = [c.strip() for c in df.columns]

        # Renommage colonnes
        rename_map = {}
        for col in df.columns:
            cu = col.strip().upper()
            if cu == "ID_CLIENT_INTERNE":
                rename_map[col] = "user_identify"
            elif cu == "CLIENT_TEL_VALIDE":
                rename_map[col] = "telephone"
        df = df.rename(columns=rename_map)

        # Si pas de colonne telephone -> on laisse la feuille comme elle est
        if "telephone" not in df.columns:
            processed_sheets[sheet_name] = df_original
            stats_rows.append({
                "Feuille": sheet_name,
                "Lignes originales": total_initial,
                "Lignes valides": total_initial,
                "Lignes non valides": 0,
                "Sans téléphone": 0,
                "Mauvais préfixe (≠ 05/06/07)": 0,
                "Doublons": 0,
                "Remarque": "Colonne CLIENT_TEL_VALIDE / telephone absente"
            })
            continue

        # Tout en texte
        df = df.astype(str)
        df["telephone"] = df["telephone"].fillna("").astype(str).str.strip()

        # Garde seulement les chiffres
        df["telephone_digits"] = df["telephone"].str.replace(r"\D", "", regex=True)

        # Option : transformer 2126xxxxxxx -> 06xxxxxxx
        def normaliser_212(x):
            if x.startswith("212") and len(x) >= 11:
                return "0" + x[3:]
            return x

        df["telephone_norm"] = df["telephone_digits"].apply(normaliser_212)

        # 3) Supprimer lignes avec téléphone vide
        mask_non_vide = df["telephone_norm"] != ""
        df_non_vide = df[mask_non_vide].copy()
        suppr_tel_vide = total_initial - len(df_non_vide)

        # 4) Garder préfixes 05 / 06 / 07
        mask_prefix = df_non_vide["telephone_norm"].str.startswith(("05", "06", "07"))
        df_prefix = df_non_vide[mask_prefix].copy()
        suppr_mauvais_prefixe = len(df_non_vide) - len(df_prefix)

        # 5) Supprimer doublons
        avant_doublons = len(df_prefix)
        df_final = df_prefix.drop_duplicates(subset="telephone_norm", keep="first").copy()
        suppr_doublons = avant_doublons - len(df_final)

        lignes_valides = len(df_final)
        lignes_invalides = total_initial - lignes_valides

        # On enlève les colonnes techniques
        df_final = df_final.drop(columns=["telephone_digits", "telephone_norm"])

        processed_sheets[sheet_name] = df_final

        stats_rows.append({
            "Feuille": sheet_name,
            "Lignes originales": total_initial,
            "Lignes valides": lignes_valides,
            "Lignes non valides": lignes_invalides,
            "Sans téléphone": suppr_tel_vide,
            "Mauvais préfixe (≠ 05/06/07)": suppr_mauvais_prefixe,
            "Doublons": suppr_doublons,
            "Remarque": ""
        })

    # Stats dans un DataFrame
    stats_df = pd.DataFrame(stats_rows)

    # Ligne TOTAL
    if not stats_df.empty:
        total_row = {
            "Feuille": "TOTAL",
            "Lignes originales": stats_df["Lignes originales"].sum(),
            "Lignes valides": stats_df["Lignes valides"].sum(),
            "Lignes non valides": stats_df["Lignes non valides"].sum(),
            "Sans téléphone": stats_df["Sans téléphone"].sum(),
            "Mauvais préfixe (≠ 05/06/07)": stats_df["Mauvais préfixe (≠ 05/06/07)"].sum(),
            "Doublons": stats_df["Doublons"].sum(),
            "Remarque": ""
        }
        stats_df = pd.concat([stats_df, pd.DataFrame([total_row])], ignore_index=True)

    return processed_sheets, stats_df


def build_combined_excel(processed_sheets: dict) -> bytes:
    output = BytesIO()
    import xlsxwriter  # s'assure que le moteur est dispo
    with pd.ExcelWriter(output, engine="xlsxwriter") as writer:
        for sheet_name, df in processed_sheets.items():
            safe_name = str(sheet_name)[:31]
            df.to_excel(writer, index=False, sheet_name=safe_name)
    output.seek(0)
    return output.getvalue()


def build_zip(processed_sheets: dict) -> bytes:
    zip_buffer = BytesIO()
    import xlsxwriter

    with zipfile.ZipFile(zip_buffer, mode="w", compression=zipfile.ZIP_DEFLATED) as zf:
        for sheet_name, (sheet_name, df) in enumerate(processed_sheets.items(), start=1):
            # Création d'un fichier Excel en mémoire pour cette feuille
            excel_buffer = BytesIO()
            with pd.ExcelWriter(excel_buffer, engine="xlsxwriter") as writer:
                safe_name = str(sheet_name)[:31]
                df.to_excel(writer, index=False, sheet_name=safe_name)
            excel_buffer.seek(0)

            # Nouveau format de nom : EQDOM_RECOUV_FIC_{INDEX}_{DATE}.xlsx
            file_name = f"EQDOM_RECOUV_FIC_{sheet_name}_{DATE_EXPORT}.xlsx"
            zf.writestr(file_name, excel_buffer.getvalue())

    zip_buffer.seek(0)
    return zip_buffer.getvalue()


def app_eqdom_marketing():
    st.title("✅ Traitement EQDOM - MARKETING")

    st.write("""
    ✅ Pour chaque feuille du fichier :
    - Colonnes en **texte (plain text)**  
    - `ID_CLIENT_INTERNE` → **user_identify**  
    - `CLIENT_TEL_VALIDE` → **telephone**  
    - Suppression des lignes sans téléphone  
    - Filtre numéros marocains **05 / 06 / 07**  
    - Suppression des doublons de téléphone  
    - Statistiques par feuille + TOTAL  

    📤 Export :
    - 1 fichier global : **EQDOM_RECOUV_FIC_SEGEMENT.xlsx**  
    - 1 fichier ZIP avec :  
      **EQDOM_RECOUV_FIC_SEGMENT1_21112025.xlsx**,  
      **EQDOM_RECOUV_FIC_SEGMENT2_21112025.xlsx**, etc.
    """)

    uploaded_file = st.file_uploader(
        "Choisir le fichier Excel (.xlsx ou .xls)",
        type=["xlsx", "xls"],
        key="eqdom_file"
    )

    if uploaded_file is not None:
        st.success(f"Fichier chargé : {uploaded_file.name}")

        if st.button("🚀 Lancer le traitement sur toutes les feuilles", key="eqdom_run"):
            # Traitement
            processed_sheets, stats_df = traiter_classeur(uploaded_file)

            # Stats
            st.subheader("📊 Statistiques par feuille")
            st.dataframe(stats_df, width='stretch')

            # Aperçu d'une feuille traitée
            st.subheader("👀 Aperçu d’une feuille traitée")
            feuille_choice = st.selectbox(
                "Choisir une feuille :",
                list(processed_sheets.keys()),
                key="eqdom_sheet_choice"
            )
            st.dataframe(processed_sheets[feuille_choice].head(), width='stretch')

            # Génération des fichiers
            combined_excel_bytes = build_combined_excel(processed_sheets)
            zip_bytes = build_zip(processed_sheets)

            st.subheader("📥 Téléchargements")

            # 1) Fichier global
            st.download_button(
                label="Télécharger le fichier global",
                data=combined_excel_bytes,
                file_name="EQDOM_RECOUV_FIC_SEGEMENT.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="eqdom_download_global"
            )

            # 2) ZIP avec 1 fichier par segment / feuille
            st.download_button(
                label="Télécharger le ZIP (1 fichier par segment)",
                data=zip_bytes,
                file_name="EQDOM_RECOUV_FIC_SEGMENTS.zip",
                mime="application/zip",
                key="eqdom_download_zip"
            )


# ==============================
# 2) APP AVT → APT (GoMobile)
#    (fichier avt2apt_app.py)
# ==============================

ALLOWED_PREFIXES = ("05", "06", "07", "08")


def normalize_phone(val: str) -> str:
    if val is None:
        return ""
    s = str(val).strip()
    if s.upper() in ("", "NAN", "NA", "NONE"):
        return ""
    s = re.sub(r"[^\d\+]", "", s)
    if s.startswith("+212"):
        s = "0" + s[4:]
    elif s.startswith("212"):
        s = "0" + s[3:]
    if s.startswith("+"):
        s = s[1:]
    if s.startswith("00212"):
        s = "0" + s[5:]
    s = re.sub(r"\D", "", s)
    if s.startswith("212"):
        s = "0" + s[3:]
    return s


def is_valid_ma_number(s: str) -> bool:
    return bool(s) and len(s) == 10 and s.startswith(ALLOWED_PREFIXES)


def read_avt(file) -> pd.DataFrame:
    # Format attendu: CLE_CONTACT|;TEL_DOM|;TEL_PRO|;TEL_GSM|;CAMPAGNE
    df = pd.read_csv(file, sep=r"\|\;", engine="python", dtype=str)
    df.columns = [c.strip().replace("\ufeff", "") for c in df.columns]
    for c in df.columns:
        df[c] = df[c].astype(str).str.strip()
    return df


def apply_business_rules(df_in: pd.DataFrame):
    """Applique normalisation + remplissage GSM + NA DOM/PRO. Renvoie df rempli + masque validité GSM Maroc."""
    cols = ["CLE_CONTACT", "TEL_DOM", "TEL_PRO", "TEL_GSM", "CAMPAGNE"]
    df = df_in.reindex(columns=cols).copy()

    for c in ["TEL_DOM", "TEL_PRO", "TEL_GSM"]:
        df[c] = df[c].apply(normalize_phone)

    def fill_gsm(row):
        if not row["TEL_GSM"]:
            if row["TEL_DOM"]:
                return row["TEL_DOM"]
            if row["TEL_PRO"]:
                return row["TEL_PRO"]
        return row["TEL_GSM"]

    df["TEL_GSM"] = df.apply(fill_gsm, axis=1)

    for c in ["TEL_DOM", "TEL_PRO"]:
        df[c] = df[c].apply(lambda x: x if x else "NA")

    for c in cols:
        df[c] = df[c].astype(str)

    valid_mask = df["TEL_GSM"].apply(is_valid_ma_number)
    return df, valid_mask


def df_to_styled_excel_bytes(df: pd.DataFrame) -> io.BytesIO:
    wb = Workbook()
    ws = wb.active
    ws.title = "DATA"
    for r in dataframe_to_rows(df, index=False, header=True):
        ws.append(r)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for r in range(1, ws.max_row + 1):
        for c in range(1, ws.max_column + 1):
            ws.cell(row=r, column=c).border = border
    # ajustement des largeurs
    for col in ws.columns:
        max_len = 10
        col_letter = col[0].column_letter
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value is not None else 0)
        ws.column_dimensions[col_letter].width = min(max_len + 2, 40)
    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return out


def app_avt_to_apt():
    st.title("📞 AVT → APT – Nettoyage & Normalisation (GoMobile)")
    st.caption(
        "Dépose tes fichiers AVT (.txt) et récupère des .xlsx prêts : "
        "normalisation, remplissage GSM, filtres Maroc, déduplication, "
        "listing des invalides, export cadrillé, et nommage CLIENT_FIC_*."
    )

    # Paramètres d'export dans la sidebar
    with st.sidebar:
        st.header("⚙️ Paramètres d'export")
        client_choice = st.selectbox(
            "Client (préfixe)",
            ["BOA", "BCP", "BMCE", "EQDOM", "AUTRE"],
            index=0,
            key="avt_client_choice"
        )
        if client_choice == "AUTRE":
            client = st.text_input(
                "Préfixe client personnalisé",
                value="CLIENT",
                key="avt_client_custom"
            ).strip() or "CLIENT"
        else:
            client = client_choice
        st.caption("Le nom final sera : CLIENT_FIC_<nom_fichier_sans_( AVT )_ni_.txt>.xlsx")

    uploaded_files = st.file_uploader(
        "Glisse-dépose un ou plusieurs fichiers AVT (.txt) au format note",
        type=["txt"],
        accept_multiple_files=True,
        key="avt_files"
    )

    if uploaded_files:
        zip_ok = io.BytesIO()
        zip_bad = io.BytesIO()

        with zipfile.ZipFile(zip_ok, "w", zipfile.ZIP_DEFLATED) as z_ok, \
             zipfile.ZipFile(zip_bad, "w", zipfile.ZIP_DEFLATED) as z_bad:

            for f in uploaded_files:
                try:
                    df_in = read_avt(f)
                    rows_in = len(df_in)
                    gsm_empty_raw = int((df_in["TEL_GSM"].astype(str).str.strip() == "").sum())

                    df_filled, valid_mask = apply_business_rules(df_in)
                    df_valid = df_filled[valid_mask].copy()
                    df_invalid = df_filled[~valid_mask].copy()

                    # Déduplication TEL_GSM (garder la 1ère occurrence)
                    before = len(df_valid)
                    df_valid = df_valid.drop_duplicates(subset=["TEL_GSM"], keep="first").copy()
                    dup_removed = int(before - len(df_valid))

                    rows_out = len(df_valid)
                    dropped_invalid = len(df_invalid)

                    # Nom de sortie : CLIENT_FIC_<base>.xlsx
                    base_name = f.name.replace("( AVT )", "").replace(".txt", "").strip()
                    out_name_ok = f"{client}_FIC_{base_name}.xlsx"
                    out_name_bad = f"{client}_FIC_{base_name}_supprimes.xlsx"

                    # ZIP des valides
                    z_ok.writestr(out_name_ok, df_to_styled_excel_bytes(df_valid).read())
                    # ZIP des invalides (si présents)
                    if not df_invalid.empty:
                        z_bad.writestr(out_name_bad, df_to_styled_excel_bytes(df_invalid).read())

                    # UI par fichier
                    st.subheader(f"🗂️ {f.name}")
                    st.markdown(f"- Lignes en entrée : **{rows_in}**")
                    st.markdown(f"- `TEL_GSM` vides avant remplissage : **{gsm_empty_raw}**")
                    st.markdown(f"- Doublons GSM supprimés : **{dup_removed}**")
                    st.markdown(f"- Lignes valides conservées : **{rows_out}**")
                    st.markdown(f"- Lignes invalides supprimées : **{dropped_invalid}**")

                    st.markdown("**Aperçu (valides)**")
                    st.dataframe(df_valid.head(10))
                    st.download_button(
                        "⬇️ Télécharger cet Excel (valides)",
                        data=df_to_styled_excel_bytes(df_valid),
                        file_name=out_name_ok,
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"ok_{f.name}"
                    )

                    if not df_invalid.empty:
                        st.markdown("🚫 **Numéros supprimés (invalides)** – aperçu")
                        st.dataframe(
                            df_invalid[["CLE_CONTACT", "TEL_DOM", "TEL_PRO", "TEL_GSM"]].head(20)
                        )
                        st.download_button(
                            "⬇️ Télécharger la liste complète des supprimés",
                            data=df_to_styled_excel_bytes(df_invalid),
                            file_name=out_name_bad,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"bad_{f.name}"
                        )
                    else:
                        st.success("✅ Aucun numéro supprimé pour ce fichier.")

                except Exception as e:
                    st.error(f"Erreur sur {f.name} : {e}")

        # Téléchargements groupés
        zip_ok.seek(0)
        zip_bad.seek(0)
        ts = datetime.now().strftime("%Y%m%d_%H%M%S")
        st.markdown("---")
        st.subheader("📦 Téléchargements groupés")
        st.download_button(
            "⬇️ ZIP – tous les Excel valides",
            data=zip_ok,
            file_name=f"APT_OK_{ts}.zip",
            mime="application/zip",
            key="avt_zip_ok"
        )
        st.download_button(
            "⬇️ ZIP – toutes les listes supprimées",
            data=zip_bad,
            file_name=f"APT_SUPPRIMES_{ts}.zip",
            mime="application/zip",
            key="avt_zip_bad"
        )

    st.markdown("---")
    st.caption(
        "Règles : remplir TEL_GSM depuis DOM/PRO si vide, DOM/PRO vides → 'NA', "
        "normaliser (+212/212→0), garder 10 chiffres démarrant par 05/06/07/08, "
        "supprimer les doublons GSM, lister les invalides. Noms: CLIENT_FIC_<base>.xlsx"
    )


# ==============================
# 3) APP Correcteur de Numéros
#    (fichier WS_MRKG_APP.py)
# ==============================

def add_zero_prefix(phone_number):
    """Ajoute un 0 au début des numéros de téléphone si nécessaire"""
    if pd.isna(phone_number):
        return phone_number

    phone_str = str(int(phone_number)) if isinstance(phone_number, float) else str(phone_number)

    # Nettoyer le numéro (enlever les espaces, etc.)
    phone_str = ''.join(filter(str.isdigit, phone_str))

    # Vérifier si le numéro ne commence pas déjà par 0 et a 9 chiffres
    if not phone_str.startswith('0') and len(phone_str) == 9:
        return '0' + phone_str
    elif len(phone_str) == 10 and phone_str.startswith('0'):
        return phone_str  # Déjà correct
    else:
        return phone_str  # Retourner tel quel si format non reconnu


def process_file(uploaded_file):
    """Traite un fichier uploadé"""
    try:
        # Lire le fichier Excel
        df = pd.read_excel(uploaded_file)

        # Vérifier si la colonne 'telephone' existe
        if 'telephone' not in df.columns:
            return None, f"❌ Colonne 'telephone' non trouvée dans {uploaded_file.name}"

        # Sauvegarder le nombre de lignes avant traitement
        original_rows = len(df)

        # Appliquer la transformation
        df['telephone'] = df['telephone'].apply(add_zero_prefix)

        # Compter les numéros modifiés (ce compteur suit la logique existante)
        modified_count = len(df[df['telephone'].astype(str).str.startswith('0')])

        # Créer un fichier en mémoire pour le téléchargement
        output = io.BytesIO()
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Feuille1')

        output.seek(0)

        return output, {
            'filename': uploaded_file.name,
            'modified_count': modified_count,
            'total_rows': original_rows,
            'df': df
        }

    except Exception as e:
        return None, f"❌ Erreur avec {uploaded_file.name}: {str(e)}"


def app_correcteur_telephone():
    st.title("📱 Correcteur de Numéros de Téléphone Pour WS MRKG")
    st.markdown("""
    Cette application ajoute automatiquement un **0** au début des numéros de téléphone 
    dans vos fichiers Excel.
    """)

    # Upload de fichiers multiples
    uploaded_files = st.file_uploader(
        "Choisissez vos fichiers Excel",
        type=['xlsx', 'xls'],
        accept_multiple_files=True,
        help="Vous pouvez sélectionner plusieurs fichiers à la fois",
        key="ws_files"
    )

    if uploaded_files:
        st.success(f"📂 {len(uploaded_files)} fichier(s) chargé(s)")

        # Bouton pour traiter tous les fichiers
        if st.button("🚀 Traiter tous les fichiers", type="primary", key="ws_run"):
            progress_bar = st.progress(0)
            status_text = st.empty()

            results = []
            errors = []

            for i, uploaded_file in enumerate(uploaded_files):
                # Mettre à jour la progression
                progress = (i + 1) / len(uploaded_files)
                progress_bar.progress(progress)
                status_text.text(f"Traitement de {uploaded_file.name}...")

                # Traiter le fichier
                output, result = process_file(uploaded_file)

                if output:
                    results.append((output, result))
                else:
                    errors.append(result)

            progress_bar.empty()
            status_text.empty()

            # Afficher les résultats
            if results:
                st.success("✅ Traitement terminé !")

                # Section de téléchargement
                st.subheader("📥 Télécharger les fichiers corrigés")

                for output, result in results:
                    filename = result['filename']
                    base_name = os.path.splitext(filename)[0]
                    new_filename = f"{base_name}_corrige.xlsx"

                    col1, col2, col3 = st.columns([3, 1, 1])

                    with col1:
                        st.write(f"**{filename}**")

                    with col2:
                        st.download_button(
                            label="📥 Télécharger",
                            data=output.getvalue(),
                            file_name=new_filename,
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"download_{filename}"
                        )

                    with col3:
                        st.info(f"{result['modified_count']}/{result['total_rows']} modifiés")

                # Aperçu des données
                st.subheader("👀 Aperçu des données corrigées")

                # Sélecteur de fichier pour l'aperçu
                file_options = {r['filename']: r for _, r in results}
                selected_file = st.selectbox(
                    "Choisir un fichier pour l'aperçu",
                    options=list(file_options.keys()),
                    key="ws_preview_choice"
                )

                if selected_file:
                    selected_result = file_options[selected_file]
                    st.caption(f"Aperçu des 10 premières lignes de {selected_file}")
                    st.dataframe(
                        selected_result['df'].head(10),
                        width='stretch'
                    )

            # Afficher les erreurs
            if errors:
                st.error("❌ Des erreurs sont survenues :")
                for error in errors:
                    st.write(error)

        # Aperçu des fichiers uploadés
        st.subheader("📋 Fichiers chargés")
        for uploaded_file in uploaded_files:
            st.write(f"• {uploaded_file.name} ({uploaded_file.size / 1024:.1f} KB)")

    else:
        # Instructions
        st.info("""
        ### 📝 Instructions :
        1. **Uploader** vos fichiers Excel contenant une colonne 'telephone'
        2. **Cliquer** sur le bouton "Traiter tous les fichiers"
        3. **Télécharger** les fichiers corrigés
        
        ### 🔧 Fonctionnalités :
        - Support multiple fichiers
        - Ajout automatique du préfixe 0
        - Aperçu des données corrigées
        - Téléchargement individuel
        """)


# ==============================
# 4) APP BOA ADAPTER
#    (portage de boa-adapter.ts en Python/Streamlit)
# ==============================

# ─── Types & Constants ───────────────────────────────────────────────────────

DELIM = '|;'

# ─── Parsers ─────────────────────────────────────────────────────────────────

def parse_out_svi(content: str) -> list[dict]:
    """Parse OUT_SVI file content. Returns list of records.
    
    Si le fichier ne contient que l'en-tête (vide), retourne une liste vide
    sans lever d'erreur pour permettre la génération d'un IN_SVI vide."""
    lines = [l.strip() for l in content.splitlines() if l.strip()]

    if len(lines) < 2:
        # Fichier vide (juste header) → retourne liste vide
        return []

    records = []
    for i in range(1, len(lines)):
        fields = lines[i].split(DELIM)
        if len(fields) < 5:
            continue
        records.append({
            'cleContact': fields[0].strip(),
            'telDom': fields[1].strip(),
            'telPro': fields[2].strip(),
            'telGsm': fields[3].strip(),
            'campagne': fields[4].strip(),
        })

    return records


def parse_out_report(content: str) -> list[dict]:
    """Parse OUT_REPORT file content. Returns list of rows.
    
    Si le fichier ne contient que l'en-tête (vide), retourne une liste vide
    sans lever d'erreur pour permettre la génération d'un IN_REPORT vide."""
    lines = [l.strip() for l in content.splitlines() if l.strip()]

    if len(lines) < 2:
        # Fichier vide (juste header) → retourne liste vide
        return []

    rows = []
    for i in range(1, len(lines)):
        fields = lines[i].split(DELIM)
        if len(fields) < 4:
            continue
        rows.append({
            'iteration': fields[0].strip(),
            'campagne': fields[1].strip(),
            'nombre': int(fields[2].strip()) if fields[2].strip().isdigit() else 0,
            'dateGeneration': fields[3].strip(),
        })

    return rows


# ─── IN_REPORT Generation ───────────────────────────────────────────────────

def generate_in_report(
    out_svi_records: list[dict],
    out_report_rows: list[dict],
    call_report: list[dict],
    audio_duration_seconds: float,
    treatment_date: str | None,
    filename_prefix: str,
    file_date_compact: str | None = None,
) -> dict:
    """Generate IN_REPORT content and stats."""
    audio_dur = audio_duration_seconds
    treatment_date = treatment_date or format_today_dotted()

    # ── Cas OUT_REPORT vide ──
    if not out_report_rows:
        # Fichier vide : générer uniquement l'en-tête, sans ligne de données
        header_fields = [
            'ITERATION', 'CAMPAGNE', 'NOMBRE', 'DATE_GENERATION',
            'DATE_TRAITEMENT_GoMobile',
            'NOMBRE_APPELS_TOTAL', 'NOMBRE_CONTACTS_TOTAL',
            'NOMBRE_CONTACTS_CONFORMES', 'NOMBRE_CONTACTS_NON_CONFORMES',
            'TAUX_CONFORMITE',
            'NOMBRE_DECROCHES', 'TAUX_DECROCHES',
            'NOMBRE_MESSAGE_ECOUTES', 'TAUX_ECOUTES',
            'NB_REECOUTE', 'TAUX_REECOUTE',
            'TOTAL_INTERACTIONS', 'TAUX_INTERACTION',
            'MINUTES_CONSOMMEES',
            'NB_BUTTON_PRESSE_1', 'NB_BUTTON_PRESSE_2', 'NB_BUTTON_PRESSE_3',
            'NB_BUTTON_PRESSE_4', 'NB_BUTTON_PRESSE_5', 'NB_BUTTON_PRESSE_6',
            'NB_BUTTON_PRESSE_7', 'NB_BUTTON_PRESSE_8', 'NB_BUTTON_PRESSE_9',
        ]

        content = DELIM.join(header_fields)

        # Filename avec la date du fichier source si disponible
        if file_date_compact:
            date_compact = file_date_compact
        else:
            date_match = re.search(r'(\d{8})', filename_prefix)
            date_compact = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')
        filename = f"{filename_prefix}_IN_REPORT_{date_compact}.txt"

        return {
            'content': content,
            'filename': filename,
            'stats': {
                'totalCalls': 0,
                'contactsTotal': 0,
                'contactsConformes': 0,
                'contactsNonConformes': 0,
                'decroches': 0,
                'ecoutes': 0,
                'reecoutes': 0,
                'interactions': 0,
                'minutesConsommees': 0,
                'buttonCounts': [0] * 9,
            },
        }

    # Reference row from OUT_REPORT (first iteration)
    ref_row = out_report_rows[0].copy()
    # Si date du fichier source fournie, remplacer la date dans ref_row
    if file_date_compact:
        ref_row['dateGeneration'] = f"{file_date_compact[:4]}.{file_date_compact[4:6]}.{file_date_compact[6:8]}"
    # NOMBRE = sum across all iterations
    total_nombre = sum(r['nombre'] for r in out_report_rows)

    # ── Compute all KPIs from GoMobile call report ──
    total_calls = 0
    decroches = 0
    ecoutes = 0
    reecoutes = 0
    total_duration_seconds = 0
    button_counts = [0] * 9  # buttons 1-9
    unique_phones = set()

    for row in call_report:
        total_calls += 1

        phone = find_field(row, 'phone', 'telephone', 'tel', 'to')
        norm = normalize_to9(phone)
        if norm:
            unique_phones.add(norm)

        call_outcome = find_field(row, 'call outcome', 'call_outcome', 'outcome')
        answered_by = find_field(row, 'answered by', 'answered_by', 'answeredby')
        duration_str = find_field(row, 'duration (s)', 'duration', 'duration_s', 'durée')
        try:
            duration = float(str(duration_str).replace(',', '.')) if duration_str and str(duration_str).lower() not in ('nan', 'nat', 'none', '') else 0.0
            if math.isnan(duration):
                duration = 0.0
        except (ValueError, TypeError):
            duration = 0.0
        dtmf = find_dtmf_input(row)

        total_duration_seconds += duration

        is_answered = (
            str(call_outcome).lower() == 'completed' and
            str(answered_by).lower() == 'human'
        )

        if is_answered:
            decroches += 1

            # Écoute: audio played completely (duration >= audio duration)
            if duration >= audio_dur:
                ecoutes += 1

            # Réécoute: audio replayed (duration >= 2x audio duration)
            if duration >= audio_dur * 2:
                reecoutes += 1

            # Button presses
            if dtmf:
                for ch in str(dtmf):
                    if ch.isdigit():
                        digit = int(ch)
                        if 1 <= digit <= 9:
                            button_counts[digit - 1] += 1

    # CONFORMES = unique valid phones from GoMobile call report
    contacts_conformes = len(unique_phones)
    contacts_total = total_nombre
    contacts_non_conformes = contacts_total - contacts_conformes

    # TOTAL_INTERACTIONS = sum of all button presses
    total_interactions = sum(button_counts)

    minutes_consommees = round(total_duration_seconds / 60) if not math.isnan(total_duration_seconds) else 0

    # ── Format percentages: 2 decimals + '%' ──
    def pct(n: float, d: float) -> str:
        return f"{(n / d * 100):.2f}%" if d > 0 else "0.00%"

    # ── Build IN_REPORT content ──
    header_fields = [
        'ITERATION', 'CAMPAGNE', 'NOMBRE', 'DATE_GENERATION',
        'DATE_TRAITEMENT_GoMobile',
        'NOMBRE_APPELS_TOTAL', 'NOMBRE_CONTACTS_TOTAL',
        'NOMBRE_CONTACTS_CONFORMES', 'NOMBRE_CONTACTS_NON_CONFORMES',
        'TAUX_CONFORMITE',
        'NOMBRE_DECROCHES', 'TAUX_DECROCHES',
        'NOMBRE_MESSAGE_ECOUTES', 'TAUX_ECOUTES',
        'NB_REECOUTE', 'TAUX_REECOUTE',
        'TOTAL_INTERACTIONS', 'TAUX_INTERACTION',
        'MINUTES_CONSOMMEES',
        'NB_BUTTON_PRESSE_1', 'NB_BUTTON_PRESSE_2', 'NB_BUTTON_PRESSE_3',
        'NB_BUTTON_PRESSE_4', 'NB_BUTTON_PRESSE_5', 'NB_BUTTON_PRESSE_6',
        'NB_BUTTON_PRESSE_7', 'NB_BUTTON_PRESSE_8', 'NB_BUTTON_PRESSE_9',
    ]

    data_fields = [
        ref_row['iteration'], ref_row['campagne'], total_nombre, ref_row['dateGeneration'],
        treatment_date,
        total_calls, total_nombre,
        contacts_conformes, contacts_non_conformes,
        pct(contacts_conformes, contacts_total),
        decroches, pct(decroches, total_calls),
        ecoutes, pct(ecoutes, decroches),
        reecoutes, pct(reecoutes, decroches),
        total_interactions, pct(total_interactions, decroches),
        minutes_consommees,
        *button_counts,
    ]

    content = DELIM.join(header_fields) + '\n' + DELIM.join(str(f) for f in data_fields)

    # Filename
    if file_date_compact:
        date_compact = file_date_compact
    else:
        date_compact = ref_row['dateGeneration'].replace('.', '')
    filename = f"{filename_prefix}_IN_REPORT_{date_compact}.txt"

    return {
        'content': content,
        'filename': filename,
        'stats': {
            'totalCalls': total_calls,
            'contactsTotal': contacts_total,
            'contactsConformes': contacts_conformes,
            'contactsNonConformes': contacts_non_conformes,
            'decroches': decroches,
            'ecoutes': ecoutes,
            'reecoutes': reecoutes,
            'interactions': total_interactions,
            'minutesConsommees': minutes_consommees,
            'buttonCounts': button_counts,
        },
    }


# ─── IN_SVI Generation ──────────────────────────────────────────────────────

def generate_in_svi(
    out_svi_records: list[dict],
    call_report: list[dict],
    date_compact: str,
    filename_prefix: str,
    file_date_compact: str | None = None,
) -> dict:
    """Generate IN_SVI content (only contacts who pressed DTMF)."""
    # Date effective à utiliser (source si disponible, sinon paramètre)
    actual_date_compact = file_date_compact if file_date_compact else date_compact

    # Si OUT_SVI vide → IN_SVI vide aussi
    if not out_svi_records:
        filename = f"{filename_prefix}_IN_SVI_{actual_date_compact}.txt"
        return {
            'content': '',
            'filename': filename,
            'recordCount': 0,
        }

    # Build lookups
    cle_to_record = {rec['cleContact']: rec for rec in out_svi_records}

    phone_to_record = {}
    for rec in out_svi_records:
        for tel in [rec['telGsm'], rec['telDom'], rec['telPro']]:
            norm = normalize_to9(tel)
            if norm:
                phone_to_record[norm] = rec

    svi_lines = []
    seen_cles = set()

    for row in call_report:
        call_outcome = find_field(row, 'call outcome', 'call_outcome', 'outcome')
        answered_by = find_field(row, 'answered by', 'answered_by', 'answeredby')
        dtmf = find_dtmf_input(row)
        phone = find_field(row, 'phone', 'telephone', 'tel', 'to')

        is_answered = (
            str(call_outcome).lower() == 'completed' and
            str(answered_by).lower() == 'human'
        )

        if not is_answered or not dtmf:
            continue

        cle = find_field(row, 'cle_contact', 'clecontact')
        # Ne garder que les contacts explicitement identifiés par cleContact dans le call_report
        # et présents dans le OUT_SVI — pas de matching par numéro de téléphone
        if not cle or cle not in cle_to_record:
            continue

        original = cle_to_record[cle]

        if cle in seen_cles:
            continue
        seen_cles.add(cle)

        line = DELIM.join([
            original['cleContact'],
            original['telDom'],
            original['telPro'],
            original['telGsm'],
            original['campagne'],
            actual_date_compact,
        ])
        svi_lines.append(line)

    filename = f"{filename_prefix}_IN_SVI_{actual_date_compact}.txt"

    return {
        'content': '\n'.join(svi_lines),
        'filename': filename,
        'recordCount': len(svi_lines),
    }


# ─── Helpers ─────────────────────────────────────────────────────────────────

def extract_filename_prefix(filename: str) -> str:
    """Extract prefix from OUT file name (e.g. 'CARTENACTIVE_OUT_SVI_20260215.txt' → 'CARTENACTIVE')."""
    base = re.sub(r'\.(txt|csv)$', '', filename, flags=re.IGNORECASE)
    match = re.match(r'^(.+?)_OUT_(?:SVI|REPORT)_', base, re.IGNORECASE)
    return match.group(1) if match else base


def normalize_to9(phone: str) -> str:
    """Normalize any phone format to last 9 digits for matching."""
    digits = re.sub(r'\D', '', str(phone))
    if len(digits) >= 9:
        return digits[-9:]
    return digits


def find_field(record: dict, *keys: str) -> str:
    """Find a value in a record by trying multiple column names (case-insensitive)."""
    for key in keys:
        norm = key.lower().strip()
        for k, v in record.items():
            if k.lower().strip() == norm and v is not None and str(v) != '':
                return str(v)
    return ''


def find_dtmf_input(record: dict) -> str:
    """Find DTMF input from dynamic 'Collect Input: Input' or similar columns."""
    for key, value in record.items():
        k = key.lower()
        if 'input' in k and 'branch' not in k and value is not None and str(value) != '':
            return str(value)
    return ''


def format_today_dotted() -> str:
    """Format today's date as YYYY.MM.DD."""
    d = datetime.now()
    return f"{d.year}.{d.month:02d}.{d.day:02d}"


# ─── Streamlit App BOA Adapter ───────────────────────────────────────────────

def app_boa_adapter():
    st.title("🔗 BOA Adapter – Génération IN_REPORT & IN_SVI")
    

    # ── Paramètres dans la sidebar ──
    with st.sidebar:
        st.header("⚙️ Paramètres GoMobile")
        audio_duration = st.number_input(
            "Durée audio (secondes)",
            min_value=1.0, max_value=250.0, value=25.0, step=1.0,
            key="boa_audio_dur"
        )
        treatment_date = st.text_input(
            "Date traitement GoMobile (YYYY.MM.DD, laisser vide = aujourd'hui)",
            value="",
            key="boa_treatment_date"
        )
        st.caption("Format: 2026.05.21")

    # ── Étape 1: Upload multi-fichiers .txt ──
    st.subheader("📁 Étape 1 : Déposer les fichiers BOA (.txt)")

    uploaded_files = st.file_uploader(
        "Glisser-déposer les fichiers .txt",
        type=["txt"],
        accept_multiple_files=True,
        key="boa_txt_upload"
    )

    out_svi_files = []
    out_report_files = []

    if uploaded_files is not None and len(uploaded_files) > 0:
        for uploaded_file in uploaded_files:
            content = uploaded_file.read().decode('utf-8', errors='replace')
            base = uploaded_file.name
            if 'OUT_SVI' in base.upper():
                out_svi_files.append({'name': base, 'content': content})
            elif 'OUT_REPORT' in base.upper():
                out_report_files.append({'name': base, 'content': content})

       

    st.subheader("📋 Étape 2 : Sélectionner les fichiers à traiter")

    selected_svi = None
    selected_report = None

    if out_svi_files and out_report_files:
        col_svi, col_report = st.columns(2)

        with col_svi:
            svi_names = [f["name"] for f in out_svi_files]
            selected_svi_name = st.selectbox(
                "BOA OUT_SVI File (.txt)",
                options=svi_names,
                key="boa_svi_select"
            )
            selected_svi = next(f for f in out_svi_files if f["name"] == selected_svi_name)

        with col_report:
            report_names = [f["name"] for f in out_report_files]
            selected_report_name = st.selectbox(
                "BOA OUT_REPORT File (.txt)",
                options=report_names,
                key="boa_report_select"
            )
            selected_report = next(f for f in out_report_files if f["name"] == selected_report_name)

    # ── Étape 3: GoMobile Call Export (CSV/Excel) ──
    st.subheader("📞 Étape 3 : GoMobile Call Export (CSV/Excel)")
    st.markdown("Fichier d'appels exporté depuis GoMobile (format CSV ou Excel).")

    call_export_file = st.file_uploader(
        "Glisser-déposer le fichier d'appels GoMobile",
        type=["csv", "xlsx", "xls"],
        accept_multiple_files=False,
        key="boa_call_export"
    )

    call_report_data = None
    if call_export_file is not None:
        try:
            if call_export_file.name.lower().endswith('.csv'):
                call_report_data = pd.read_csv(call_export_file, dtype=str).to_dict('records')
            else:
                call_report_data = pd.read_excel(call_export_file, dtype=str).to_dict('records')
            st.success(f"✅ {len(call_report_data)} appels chargés depuis {call_export_file.name}")
        except Exception as e:
            st.error(f"❌ Erreur lecture fichier d'appels : {e}")

    # ── Étape 4: Traitement ──
    st.subheader("🚀 Étape 4 : Génération des fichiers IN")

    # ── Détection fichiers vides ──
    svi_empty = False
    report_empty = False
    if selected_svi is not None:
        try:
            svi_test = parse_out_svi(selected_svi['content'])
            svi_empty = len(svi_test) == 0
        except Exception:
            svi_empty = True
    if selected_report is not None:
        try:
            report_test = parse_out_report(selected_report['content'])
            report_empty = len(report_test) == 0
        except Exception:
            report_empty = True

    both_empty = svi_empty and report_empty

    can_process = (
        selected_svi is not None and
        selected_report is not None and
        (call_report_data is not None or both_empty)
    )

    

    if both_empty:
        st.info("ℹ️ Les fichiers OUT_SVI et OUT_REPORT sont vides — la génération produira des fichiers IN vides (headers seuls).")

    if st.button("▶️ Lancer la génération IN_REPORT + IN_SVI", disabled=not can_process, key="boa_run"):
        try:
            with st.spinner("Traitement en cours..."):
                # Parse
                out_svi_records = parse_out_svi(selected_svi['content'])
                out_report_rows = parse_out_report(selected_report['content'])

                # Extract prefix
                filename_prefix = extract_filename_prefix(selected_svi['name'])

                # Extract date from source filename (e.g. WELCOME_OUT_SVI_20260521.txt → 20260521)
                source_date_match = re.search(r'(\d{8})', selected_svi['name'])
                source_date_compact = source_date_match.group(1) if source_date_match else None

                # Validate treatment date
                treatment_date_val = treatment_date.strip() if treatment_date else None
                if treatment_date_val and not re.match(r'^\d{4}\.\d{2}\.\d{2}$', treatment_date_val):
                    st.warning("Format date invalide, utilisation de la date du jour.")
                    treatment_date_val = None

                # Si pas de Call Export mais fichiers vides → utiliser liste vide
                call_data = call_report_data if call_report_data is not None else []

                # Generate IN_REPORT
                in_report = generate_in_report(
                    out_svi_records,
                    out_report_rows,
                    call_data,
                    audio_duration,
                    treatment_date_val,
                    filename_prefix,
                    source_date_compact,
                )

                # Generate IN_SVI
                if out_report_rows:
                    date_compact = out_report_rows[0]['dateGeneration'].replace('.', '')
                else:
                    date_match = re.search(r'(\d{8})', filename_prefix)
                    date_compact = date_match.group(1) if date_match else datetime.now().strftime('%Y%m%d')
                in_svi = generate_in_svi(
                    out_svi_records,
                    call_data,
                    date_compact,
                    filename_prefix,
                    source_date_compact,
                )

            # ── Affichage des résultats ──
            st.success("✅ Génération terminée avec succès !")

            # Stats
            st.subheader("📊 Statistiques IN_REPORT")
            stats = in_report['stats']

            c1, c2, c3, c4 = st.columns(4)
            c1.metric("Appels totaux", stats['totalCalls'])
            c2.metric("Contacts total", stats['contactsTotal'])
            c3.metric("Contacts conformes", stats['contactsConformes'])
            c4.metric("Contacts non conformes", stats['contactsNonConformes'])

            c5, c6, c7, c8 = st.columns(4)
            c5.metric("Décrochés", stats['decroches'])
            c6.metric("Écoutes", stats['ecoutes'])
            c7.metric("Réécoutes", stats['reecoutes'])
            c8.metric("Interactions (DTMF)", stats['interactions'])

            st.metric("Minutes consommées", stats['minutesConsommees'])

            # Buttons pressés
            st.subheader("🔘 Compteurs de touches DTMF")
            btn_cols = st.columns(9)
            for i, count in enumerate(stats['buttonCounts']):
                btn_cols[i].metric(f"Touche {i+1}", count)

            # ── Téléchargements ──
            st.subheader("📥 Téléchargements")

            col_dl1, col_dl2 = st.columns(2)

            with col_dl1:
                st.download_button(
                    label=f"⬇️ Télécharger {in_report['filename']}",
                    data=in_report['content'].encode('utf-8'),
                    file_name=in_report['filename'],
                    mime="text/plain",
                    key="boa_dl_in_report"
                )

            with col_dl2:
                st.download_button(
                    label=f"⬇️ Télécharger {in_svi['filename']}",
                    data=in_svi['content'].encode('utf-8'),
                    file_name=in_svi['filename'],
                    mime="text/plain",
                    key="boa_dl_in_svi"
                )

            # ── Aperçus ──
            with st.expander("👀 Aperçu IN_REPORT"):
                st.text(in_report['content'])

            with st.expander(f"👀 Aperçu IN_SVI ({in_svi['recordCount']} lignes)"):
                if in_svi['recordCount'] > 0:
                    preview_lines = in_svi['content'].split('\n')[:10]
                    st.text('\n'.join(preview_lines))
                    if in_svi['recordCount'] > 10:
                        st.caption(f"... et {in_svi['recordCount'] - 10} lignes supplémentaires")
                else:
                    st.info("Aucun contact n'a pressé de touche DTMF.")

        except Exception as e:
            st.error(f"❌ Erreur lors du traitement : {e}")
            st.exception(e)

    st.markdown("---")
    st.caption(
        "Logique : parse OUT_SVI/OUT_REPORT → croisement avec Call Export GoMobile → "
        "KPIs (décrochés, écoutes, réécoutes, DTMF) → génération IN_REPORT + IN_SVI."
    )


# ==============================
# ROUTEUR PRINCIPAL (MENU)
# ==============================

st.sidebar.title("🛠️ Menu des outils")
app_choice = st.sidebar.radio(
    "Choisir l'application :",
    (
        "EQDOM_MARKETING",
        "BOA_MARKETING",
        "WS_MRKG_APP_ADD_0",
        "BOA_REPORT_GENERATOR",
    ),
    key="main_app_choice"
)

if app_choice == "EQDOM_MARKETING":
    app_eqdom_marketing()
elif app_choice == "BOA_MARKETING":
    app_avt_to_apt()
elif app_choice == "BOA_REPORT_GENERATOR":
    app_boa_adapter()
else:
    app_correcteur_telephone()
